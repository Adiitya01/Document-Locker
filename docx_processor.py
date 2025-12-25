import zipfile
from lxml import etree
import shutil
import os
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
import xlrd
from xlutils.copy import copy as xlcopy

# ---------------- CONFIG ----------------
TEMP_BASE_DIR = "temp"
os.makedirs(TEMP_BASE_DIR, exist_ok=True)

NAMESPACE = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
}

# ---------------- CORE LOGIC ----------------
def is_cell_empty(cell):
    """Check if a table cell is empty or contains only whitespace"""
    # Get all text content in the cell
    texts = cell.xpath(".//w:t", namespaces=NAMESPACE)
    
    if not texts:
        return True
    
    # Check if all text nodes are empty or whitespace
    all_text = "".join([t.text or "" for t in texts])
    return all_text.strip() == ""

def wrap_cell_in_content_control(cell, cell_id):
    """Wrap the entire cell content in a content control"""
    # Get all paragraphs in the cell
    paragraphs = cell.xpath(".//w:p", namespaces=NAMESPACE)
    
    if not paragraphs:
        return
    
    # Create content control to wrap the first paragraph
    sdt = etree.Element("{%s}sdt" % NAMESPACE["w"])
    
    # Properties
    sdtPr = etree.SubElement(sdt, "{%s}sdtPr" % NAMESPACE["w"])
    
    # Unique ID
    sdt_id = etree.SubElement(sdtPr, "{%s}id" % NAMESPACE["w"])
    sdt_id.set("{%s}val" % NAMESPACE["w"], str(cell_id))
    
    # Plain text control
    etree.SubElement(sdtPr, "{%s}text" % NAMESPACE["w"])
    
    # Content
    sdtContent = etree.SubElement(sdt, "{%s}sdtContent" % NAMESPACE["w"])
    
    # Move the first paragraph into the content control
    first_para = paragraphs[0]
    parent = first_para.getparent()
    
    # Get the index of the first paragraph
    para_index = list(parent).index(first_para)
    
    # Remove first paragraph from cell
    parent.remove(first_para)
    
    # Add paragraph to content control
    sdtContent.append(first_para)
    
    # Insert content control at the original position
    parent.insert(para_index, sdt)

def protect_docx(input_path: str, output_path: str, password: str = None):
    temp_dir = os.path.join(TEMP_BASE_DIR, f"docx_{uuid.uuid4().hex}")
    os.makedirs(temp_dir)

    try:
        # 1. Unzip DOCX
        with zipfile.ZipFile(input_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        document_xml = os.path.join(temp_dir, "word/document.xml")
        tree = etree.parse(document_xml)
        root = tree.getroot()

        # 2. Find all table cells and wrap empty ones in content controls
        empty_cells = 0
        filled_cells = 0
        cell_id = 1000000
        
        for table in root.xpath("//w:tbl", namespaces=NAMESPACE):
            for row in table.xpath(".//w:tr", namespaces=NAMESPACE):
                for cell in row.xpath(".//w:tc", namespaces=NAMESPACE):
                    if is_cell_empty(cell):
                        wrap_cell_in_content_control(cell, cell_id)
                        empty_cells += 1
                        cell_id += 1
                    else:
                        filled_cells += 1

        # 3. Save document.xml
        tree.write(document_xml, xml_declaration=True, encoding="utf-8", standalone="yes")

        # 4. Enable document protection (Forms only)
        settings_xml = os.path.join(temp_dir, "word/settings.xml")
        
        if os.path.exists(settings_xml):
            settings_tree = etree.parse(settings_xml)
            settings_root = settings_tree.getroot()
        else:
            # Create settings.xml if it doesn't exist
            settings_root = etree.Element(
                "{%s}settings" % NAMESPACE["w"],
                nsmap={"w": NAMESPACE["w"]}
            )
            settings_tree = etree.ElementTree(settings_root)

        # Remove existing protection
        for el in settings_root.xpath("//w:documentProtection", namespaces=NAMESPACE):
            settings_root.remove(el)

        # Add protection
        protection = etree.Element("{%s}documentProtection" % NAMESPACE["w"])
        protection.set("{%s}edit" % NAMESPACE["w"], "forms")
        protection.set("{%s}enforcement" % NAMESPACE["w"], "1")

        settings_root.insert(0, protection)
        settings_tree.write(settings_xml, xml_declaration=True, encoding="utf-8", standalone="yes")

        # 5. Zip back to DOCX
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for foldername, _, filenames in os.walk(temp_dir):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = file_path.replace(temp_dir + os.sep, "")
                    zipf.write(file_path, arcname)

        return empty_cells, filled_cells

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

def protect_xlsx(input_path: str, output_path: str):
    """Lock filled cells in Excel, keep empty cells editable"""
    try:
        file_ext = input_path.split('.')[-1].lower()
        
        # Handle .xls format (convert to .xlsx)
        if file_ext == "xls":
            # Read .xls file with xlrd
            xlrd_book = xlrd.open_workbook(input_path)
            xlrd_sheet = xlrd_book.sheet_by_index(0)
            
            # Create new openpyxl workbook
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            
            empty_cells = 0
            filled_cells = 0
            
            # Copy data from xlrd to openpyxl
            for row_idx in range(xlrd_sheet.nrows):
                for col_idx in range(xlrd_sheet.ncols):
                    cell_value = xlrd_sheet.cell_value(row_idx, col_idx)
                    openpyxl_cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    
                    if cell_value is None or str(cell_value).strip() == "":
                        empty_cells += 1
                        openpyxl_cell.protection = Protection(locked=False, hidden=False)
                    else:
                        filled_cells += 1
                        openpyxl_cell.protection = Protection(locked=True, hidden=False)
            
            # Enable sheet protection
            ws.protection = SheetProtection(sheet=True, password=None)
            
            wb.save(output_path)
        
        # Handle .xlsx format
        else:
            wb = load_workbook(input_path)
            ws = wb.active
            
            empty_cells = 0
            filled_cells = 0
            
            # Iterate through all cells with values
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        empty_cells += 1
                        # Unlock empty cells
                        cell.protection = Protection(locked=False, hidden=False)
                    else:
                        # Lock filled cells
                        cell.protection = Protection(locked=True, hidden=False)
                        filled_cells += 1
            
            # Enable sheet protection
            ws.protection = SheetProtection(sheet=True, password=None)
            
            wb.save(output_path)
        
        return empty_cells, filled_cells
        
    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")

